from unittest import result
from django.http import HttpResponse
from io import BytesIO
import os
from django.template.loader import get_template
import openpyxl
from xhtml2pdf import pisa
import openpyxl
from openpyxl import load_workbook

def render_to_pdf(template_src, context_dict={}):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')

    return None
Ruta_excel = "D:/10mo/Tesis2/Postest.xlsx"
def iniciar_excel():
    if not os.path.exists(Ruta_excel):
        workbook = openpyxl.Workbook()
        hoja=workbook.active
        hoja.title = "Indicador 1"

        hoja.append(["Fecha", "Tiempo (segundos)"])
        workbook.save(Ruta_excel)
def actualizar_excel(fecha, tiempo):
    workbook = load_workbook(Ruta_excel)
    hoja = workbook.active

    hoja.append([fecha, tiempo])
    workbook.save(Ruta_excel)